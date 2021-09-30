# download openpyxl from pypi.org
# NOTE: Ensure excel file is in same directory/folder as python file

import openpyxl as xl


# def customer_discount(filename, sheet_name):
#     wb = xl.load_workbook(filename)
#     sheet = wb[sheet_name]
#     for row in range(2, sheet.max_row + 1):
#         cell_1 = sheet.cell(row, 2)
#         cell_2 = sheet.cell(row, 3)
#         nxt_disc = cell_1.value * cell_2.value / 100  # calculating next discount for customer based on age (cell_1.value) and expenditure (cell_2.value)
#         nxt_dis_cell = sheet.cell(row, 4)
#         nxt_dis_cell.value = nxt_disc

#     wb.save(filename)


# function adapted for heart2.xlsx
num = 1

def process_heart_xl(filename, sheet_name):

    wb = xl.load_workbook(filename)
    sheet = wb[sheet_name]

    global num

    for row in range(2, sheet.max_row + 1):
        cell_age = sheet.cell(row, 1)
        cell_trestbps = sheet.cell(row, 4)
        cell_thalach = sheet.cell(row, 8)
        cell_chol = sheet.cell(row, 15)
        # print(type(cell_trestbps))
        if (cell_age.value > 45 and cell_thalach.value > 160 and
            cell_chol.value > 250 and cell_trestbps.value > 130): #tresbps = resting blood pressure || thalach = maximum heart rate achieved
            disease = 1
            print('Patient '+ str(num) +' has disease = 1')
        else:
            print('---')
            disease = 0
        pred_cell = sheet.cell(row, 5)
        pred_cell.value = disease
        num+= 1
    wb.save(filename)

# Example of how to use function in script

# from excel_functions import process_heart_xl
#
# process_heart_xl('heart2.xlsx', 'heart2')

process_heart_xl('DataTest.xlsx', 'DataTest')

def conv_target_xl(filename, sheet_name):

    wb = xl.load_workbook(filename)
    sheet = wb[sheet_name]

    for row in range(2, sheet.max_row + 1):

        cell_target = sheet.cell(row, 14)
        if cell_target.value == 2:
            cell_target.value = 1

    wb.save(filename)
